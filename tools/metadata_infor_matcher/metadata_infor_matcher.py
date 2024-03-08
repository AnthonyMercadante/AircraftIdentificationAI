"""
File Name: Metadata Information Matcher Script
Desc: This script matches metadata from the planes Excel sheet to the photos
      We are able to identify each plane by its metadata, as well a description on what each image is

      Done:
      - Display photo information instead of metadata //done
      - Fix suffix logic // done
      - remove duplicates from appearing // done??? double check but should be fixed
      - attach description to each modified photo (expand array for description section) // done

      To-Do:
      - Dynamically generate sheet names, instead of hardcoding //insert your modifications nam
      - fix sheet formatting on dates (ensure its all the same) - should be done on excel (should look like 1984.26.51.a-gg)
      - pickle the metadata dates // ill do that at the end - adam
      - Create a dict for most common words (will help us see what plane names appear frequently)
      - bonus: store data into it's own individual text file (possible option)
      - ensure suffix logic works for both upper and lowercase
      - remove false positives (i.e. 1984.26.8 giving the same description for 1984.26.81 1984.26.82 ...)


Authors: Nguyen Quoc Nam Tran, Adam Calleja
IDs: 000876813, 000862779
School: Mohawk College
Date Created: February 17th, 2024
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import re
# import spacy
from collections import Counter



def get_suffix_list(suffix):
    """
    This function creates a list of all the suffix values from the starting to ending value The function will
    check to see how long the suffix is (i.e. a-g or a-aa) and handle accordingly The function will ensure that all
    values from the start to the end are displayed in a list Used when dealing with metadata such as 1984.19.47.a-s
    which signifies that there are photos with the same date with suffixes from a-s

    Params:
    - suffix (string): Suffix value used to signify different photos from the same snapshot date

    Return:
    - suffix_list (list): List of all the suffix values from the start to the end of the suffix
    """

    # No suffix located
    if not suffix or '-' not in suffix:
        return [suffix]

    # Find the first and last occurrence of '-' to determine the actual range
    first_dash = suffix.find('-')
    last_dash = suffix.rfind('-')

    # If there's more than one '-', handle specially, otherwise proceed normally
    if first_dash != last_dash:
        # Complex range: take everything up to the first '-' as start, and after the last '-' as end
        start, end = suffix[:first_dash], suffix[last_dash + 1:]
    else:
        # Simple range: split normally
        start, end = suffix.split('-')

    suffix_list = []
    start_index = ord(start[0])
    end_index = ord(end[0]) if end else start_index  # Handle single character or simple range

    if len(start) == 1 and (len(end) == 1 or not end):
        # Generate list for single character ranges
        suffix_list = [chr(i) for i in range(start_index, end_index + 1)]
    else:
        # Handle more complex ranges (e.g., a-aa, a-zz)
        current_char = start
        while current_char != end:
            suffix_list.append(current_char)
            # Logic to increment characters ('a' -> 'b', 'z' -> 'aa', etc.)
            if len(current_char) == 1:
                if current_char < 'z':
                    current_char = chr(ord(current_char) + 1)
                else:
                    current_char = 'aa'
            else:
                if current_char[1] < 'z':
                    current_char = current_char[0] + chr(ord(current_char[1]) + 1)
                else:
                    if current_char[0] < 'z':
                        current_char = chr(ord(current_char[0]) + 1) + 'a'
                    else:
                        break  # End of range
        if end:  # Ensure the end of the range is included
            suffix_list.append(end)

    return suffix_list


def normalize_date_format(date_str):
    """
    This function normalizes the data so that it is processed in uniform
    Ensures that the dates are all separated by .'s, and that we break apart the date with the suffix
    If no suffix exists for the date, we return back a blank suffix

    Params:
    date_str (string): Metadata String that we want to format

    Return:
        Tuple:
        standardized_date_str (string): Updated and formatted date
        suffix (string): Suffix that follows the date, blank or letter value
    """
    standardized_date_str = date_str.replace('-', '.').replace(' ', '.')

    # Split the string into parts based on '.'
    parts = standardized_date_str.split('.')

    # Check if there's more than three parts indicating a suffix exists
    if len(parts) > 3:
        numeric_part = '.'.join(parts[:3])  # Join the first three parts as the date
        suffix = '.'.join(parts[3:])  # Join the remaining parts as the suffix
    else:
        numeric_part = standardized_date_str  # The date without a suffix
        suffix = ''  # No suffix present

    return numeric_part, suffix


def create_dates_array(metadata_dates):
    """
    This function creates an array with all the dates and ensures that all the suffix names are appended to the date
    Appending the suffix value to the date allows us to search for all images within a suffix range

    Params:
        metadata_dates (list of tuple): Contains a date, and a description

    Returns:
        all_dates_with_descriptions (dict): Uses dates as keys and descriptions as values
    """
    all_dates_with_descriptions = {}
    for date_str, description in metadata_dates:
        base_date, suffix_range = normalize_date_format(date_str)
        suffixes = get_suffix_list(suffix_range.replace('.', '-')) if suffix_range else ['']
        for suffix in suffixes:
            full_date = f"{base_date}{suffix}"
            all_dates_with_descriptions[full_date] = description
    return all_dates_with_descriptions


def find_matches(generated_dates_with_descriptions, image_folder):
    """
    Used to locate matches between metadata dates and image filenames
    If a filename is located with metadata, then the image file and description are returned
    If a filename is located without metadata, then the image file with the word "None" are returned signifying no description
    Ensures that filenames are not output more than once

    Params:
        generated_dates_with_descriptions (dict): Dates with their descriptions
        image_folder(string): Path to the folder with the contained image files

    Returns:
        matched_photos_with_description (list): List of the image file names with its description

    """
    matched_photos_with_descriptions = []
    matched_filenames = set()  # Track which filenames have already been processed

    if not os.path.exists(image_folder):
        print(f"Image folder does not exist: {image_folder}")
        return matched_photos_with_descriptions

    image_files = [f for f in os.listdir(image_folder) if os.path.isfile(os.path.join(image_folder, f))]

    for date, description in generated_dates_with_descriptions.items():
        for filename in image_files:
            # Check if the date is in the filename and the filename hasn't been processed yet
            if date in filename.replace('-', '.').lower() and filename not in matched_filenames:
                matched_photos_with_descriptions.append((filename, description))
                matched_filenames.add(filename)  # Mark this filename as processed

    # Filenames that haven't matched any description from metadata
    for filename in image_files:
        if filename not in matched_filenames:
            matched_photos_with_descriptions.append((filename, "None"))

    return matched_photos_with_descriptions


def load_metadata_from_excel(excel_file_path, sheet_name):
    """
    Loads the metadata from the Excel sheet, extracting the date and description

    Params:
        excel_file_path (string): Path to the Excel file
        sheet_name (str): Name of sheet to be processed

    Returns:
        metadata (list): List of tuples containing the date and the description
    """
    workbook = load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    metadata = []
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True):
        date, description = row
        if date:
            metadata.append((str(date), description))
    return metadata

def process_sheet(sheet_name, base_path):
    """
    Process a sheet from the Excel file, matching metadata to filenames
    Output matches between metadata and filenames

    Params:
        sheet_name (string): Name of the sheet to process
        base_path (string): Directory path where Excel and image files are located

    """
    extended_base_path = os.path.join(base_path, 'Anthony_Nish_Adam_Nguyen_Group', 'NAFMC')
    excel_file = os.path.join(extended_base_path, 'NAFMC Photographic Collection.xlsx')

    metadata = load_metadata_from_excel(excel_file, sheet_name)

    generated_dates_with_descriptions = {}
    for date_str, description in metadata:
        base_date, suffix_range = normalize_date_format(date_str)
        suffixes = get_suffix_list(suffix_range.replace('.', '-')) if suffix_range else ['']
        for suffix in suffixes:
            generated_date = f"{base_date}{suffix}"
            generated_dates_with_descriptions[generated_date] = description

       

    image_folder = os.path.join(extended_base_path, sheet_name.replace('-', ' '))
    matched_photos_with_descriptions = find_matches(generated_dates_with_descriptions, image_folder)

    if matched_photos_with_descriptions:
        print(f"Matches found in sheet '{sheet_name}':")
        for photo, description in matched_photos_with_descriptions:
            print(f"{photo} {description}")
    else:
        print(f"No matches found in sheet '{sheet_name}'.")


# Load spaCy's pre-trained English model
# nlp = spacy.load("en_core_web_sm")
#Processing Test
base_path = os.path.dirname(os.path.abspath(__file__));
excel_file = os.path.join(base_path, 'NAFMC Photographic Collection.xlsx')
workbook = load_workbook(excel_file)
sheet_names = workbook.sheetnames


# Process each sheet
for sheet_name in sheet_names:
    print(f"\nProcessing sheet: {sheet_name}")
    process_sheet(sheet_name, base_path)


excel_file_aircraft_model = os.path.join(base_path, 'List_of_aircraft_of_Canada_refs_wikipedia.xlsx')

workbook_aircraft_model = load_workbook(excel_file_aircraft_model)


# Create a dictionary to store descriptions for each sheet
all_descriptions = {}

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    #the "Description" column is in column D (column index 4)
    description_column = sheet['D']

    # Extract all descriptions from the column
    descriptions = [str(cell.value).lower() for cell in description_column[1:] if cell.value]

    # Store the descriptions in the dictionary with sheet_name as the key
    all_descriptions[sheet_name] = descriptions

# Create a dictionary of designators and their associated names from the same sheet as aircraft models
aircraft_models_sheet = workbook_aircraft_model[workbook_aircraft_model.sheetnames[0]]
designators_sheet = workbook_aircraft_model[workbook_aircraft_model.sheetnames[0]]

aircraft_models = [str(model.value).lower() for model in aircraft_models_sheet['A'][1:] if model.value]
designator_name_dict = {str(desig.value).lower(): str(name.value) for desig, name in zip(designators_sheet['B'][1:], aircraft_models_sheet['A'][1:]) if desig.value}

# convert to lowercase
lowercase_designators = [item.lower() for sublist in designators_sheet['B'][1:] for item in str(sublist.value).split()]

print("\nAll Designators:")
print(lowercase_designators)

# Iterate through all sheets in the all_descriptions dictionary
for sheet_name, descriptions in all_descriptions.items():
    print(f"\nSheet: {sheet_name}")

    for description in descriptions:
        description_lower = description.lower()

        # Check for matches in the aircraft model list from wikipedia
        matches_aircraft_model = [model for model in aircraft_models if model in description_lower]

        # Check for matches in the designator list from wikipedia
        matches_designator = [desig for desig in lowercase_designators if desig in description_lower]

        # Print the description along with matching aircraft models and designators
        if matches_aircraft_model or matches_designator:
            print(f"\nDescription: {description}")
            if matches_aircraft_model:
                print(f"Matches - Aircraft Models: {matches_aircraft_model}")
            if matches_designator:
                matching_names = [designator_name_dict[desig] for desig in matches_designator if desig in designator_name_dict]
                print(f"Matches - Designators: {matches_designator}")
                print(f"Matching Names: {matching_names}")


workbook.close()
workbook_aircraft_model.close()



